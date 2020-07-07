import tkinter as tk
from tkinter import filedialog                #File dialogs help you open, save files or directories.
import pandas as pd
from openpyxl import *
from tkinter.messagebox import showinfo
def open():
    path=filedialog.askopenfilename()           #ask for the file location,locate where your file is placed
    df = pd.read_excel(path)                    #read the file path
    print(df)                                   #print the data in file
def save():
    fname = entry.get()
    lname = entry1.get()
    age = entry2.get()
    email = entry3.get()
    mobile = entry4.get()
    city = entry5.get()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "First Name"
    ws['B1'] = "Last Name"
    ws['C1'] = "Age"
    ws['D1'] = "Email"
    ws['E1'] = "Mobile"
    ws['F1'] = "City"
    ws['A2']=fname
    ws['B2']=lname
    ws['C2']=age
    ws['D2']=email
    ws['E2']=mobile
    ws['F2']=city
    wb.save(r'D:\Mini codes\Registration\reg1.xlsx')     #use reg1.xlsx to store new data
    showinfo("Saved","Your Entry has been saved")
    file1 = pd.read_excel("reg.xlsx")                    #use reg.xlsx to store all the data to avoid rewriting
    file2 = pd.read_excel("reg1.xlsx")
    all = [file1, file2]                                 #define the sequence of data to be stored
    append = pd.concat(all)                              #concatenate the files
    append.to_excel("reg.xlsx", index=False)             #write the result to reg.xlsx
def clear():
    entry.delete(0, tk.END)
    entry1.delete(0, tk.END)
    entry2.delete(0, tk.END)
    entry3.delete(0, tk.END)
    entry4.delete(0, tk.END)
    entry5.delete(0, tk.END)
win=tk.Tk()
win.title("Registration Form")
label = tk.Label(win, text="First Name:")
label.grid(row=0, column=0, padx=8, pady=8)
entry = tk.Entry(win)
entry.grid(row=0, column=1, padx=8, pady=8)
label1 = tk.Label(win, text="Last Name:")
label1.grid(row=1, column=0, padx=8, pady=8)
entry1 = tk.Entry(win)
entry1.grid(row=1, column=1, padx=8, pady=8)
label2 = tk.Label(win, text="Age:")
label2.grid(row=2, column=0, padx=8, pady=8)
entry2 = tk.Entry(win)
entry2.grid(row=2, column=1, padx=8, pady=8)
label3 = tk.Label(win, text="Email:")
label3.grid(row=3, column=0, padx=8, pady=8)
entry3 = tk.Entry(win)
entry3.grid(row=3, column=1, padx=8, pady=8)
label4 = tk.Label(win, text="Mobile Number:")
label4.grid(row=4, column=0, padx=8, pady=8)
entry4 = tk.Entry(win)
entry4.grid(row=4, column=1, padx=8, pady=8)
label5 = tk.Label(win, text="City:")
label5.grid(row=5, column=0, padx=8, pady=8)
entry5 = tk.Entry(win)
entry5.grid(row=5, column=1, padx=8, pady=8)
button = tk.Button(win, text="Register",command=save)
button.grid(row=6, column=0, padx=8, pady=8)
button1 = tk.Button(win, text="Clear",command=clear)
button1.grid(row=6, column=1, padx=8, pady=8)
button2=tk.Button(win,text="Load Excel",command=open)
button2.grid(row=6, column=2, padx=8, pady=8)
win.mainloop()
